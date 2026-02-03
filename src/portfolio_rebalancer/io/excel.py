"""Excel I/O handler for portfolio data."""

from __future__ import annotations

from typing import List

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "openpyxl is required for Excel support. Install with: pip install openpyxl"
    )

from portfolio_rebalancer.models.asset import Asset
from portfolio_rebalancer.models.portfolio import Portfolio
from portfolio_rebalancer.models.result import RebalancingResult


class ExcelIO:
    """Handler for reading and writing portfolio data from/to Excel files.
    
    Excel file format for input:
    - Row 1: "Cash Available" label in A1, value in B1
    - Row 2: Empty (separator)
    - Row 3: Headers
    - Row 4+: Asset data
    
    Columns:
    - Column A: Symbol (ticker)
    - Column B: Quantity (shares owned)
    - Column C: Price (current price)
    - Column D: Avg Cost (average purchase price)
    - Column E: Tax Rate (decimal, e.g., 0.26)
    - Column F: Target Weight (decimal, e.g., 0.60)
    - Column G: Commission Buy Fixed (€, e.g., 2.50)
    - Column H: Commission Buy Percent (decimal, e.g., 0.001 for 0.1%)
    - Column I: Commission Buy Min (€, e.g., 1.00)
    - Column J: Commission Buy Max (€, e.g., 10.00)
    - Column K: Commission Sell Fixed (€, e.g., 2.50)
    - Column L: Commission Sell Percent (decimal, e.g., 0.001 for 0.1%)
    - Column M: Commission Sell Min (€, e.g., 1.00)
    - Column N: Commission Sell Max (€, e.g., 10.00)
    """
    
    def read_portfolio(self, filepath: str, sheet_name: str = None) -> Portfolio:
        """Read portfolio from Excel file.
        
        Args:
            filepath: Path to Excel file
            sheet_name: Name of sheet to read (uses active sheet if None)
            
        Returns:
            Portfolio object with assets from Excel
            
        Raises:
            ValueError: If Excel file is invalid or data is malformed
        """
        try:
            workbook = openpyxl.load_workbook(filepath)
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {e}")
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # Read cash available from B1 (A1 should be "Cash Available")
        cash_available = 0.0
        if sheet['A1'].value and 'cash' in str(sheet['A1'].value).lower():
            try:
                cash_value = sheet['B1'].value
                if cash_value is not None:
                    cash_available = float(cash_value)
            except (ValueError, TypeError):
                # If cash_available is not a valid number, default to 0.0
                pass
        
        assets = []
        
        # Skip header row (row 3) and cash info rows (1-2)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            # Skip empty rows
            if not any(row):
                continue
            
            # Support old format (6 columns) and new format (14 columns)
            if len(row) < 6:
                raise ValueError(
                    f"Row {row_idx} has insufficient columns. Expected at least 6 columns: "
                    "Symbol, Quantity, Price, Avg Cost, Tax Rate, Target Weight "
                    "(+ optional commission columns)"
                )
            
            # Basic asset fields
            symbol, quantity, price, avg_cost, tax_rate, target_weight = row[:6]
            
            # Commission fields (default to 0.0 if missing or None)
            commission_buy_fixed = row[6] if len(row) > 6 and row[6] is not None else 0.0
            commission_buy_percent = row[7] if len(row) > 7 and row[7] is not None else 0.0
            commission_buy_min = row[8] if len(row) > 8 and row[8] is not None else 0.0
            commission_buy_max = row[9] if len(row) > 9 and row[9] is not None else 0.0
            
            commission_sell_fixed = row[10] if len(row) > 10 and row[10] is not None else 0.0
            commission_sell_percent = row[11] if len(row) > 11 and row[11] is not None else 0.0
            commission_sell_min = row[12] if len(row) > 12 and row[12] is not None else 0.0
            commission_sell_max = row[13] if len(row) > 13 and row[13] is not None else 0.0
            
            # Validate data
            try:
                symbol = str(symbol).strip()
                quantity = float(quantity)
                price = float(price)
                avg_cost = float(avg_cost)
                tax_rate = float(tax_rate)
                target_weight = float(target_weight)
                
                # Commission values
                commission_buy_fixed = float(commission_buy_fixed)
                commission_buy_percent = float(commission_buy_percent)
                commission_buy_min = float(commission_buy_min)
                commission_buy_max = float(commission_buy_max)
                
                commission_sell_fixed = float(commission_sell_fixed)
                commission_sell_percent = float(commission_sell_percent)
                commission_sell_min = float(commission_sell_min)
                commission_sell_max = float(commission_sell_max)
            except (ValueError, TypeError) as e:
                raise ValueError(
                    f"Invalid data in row {row_idx}: {e}. "
                    "Ensure all numeric columns contain valid numbers."
                )
            
            asset = Asset(
                symbol=symbol,
                quantity=quantity,
                price=price,
                avg_cost=avg_cost,
                tax_rate=tax_rate,
                target_weight=target_weight,
                commission_buy_fixed=commission_buy_fixed,
                commission_buy_percent=commission_buy_percent,
                commission_buy_min=commission_buy_min,
                commission_buy_max=commission_buy_max,
                commission_sell_fixed=commission_sell_fixed,
                commission_sell_percent=commission_sell_percent,
                commission_sell_min=commission_sell_min,
                commission_sell_max=commission_sell_max,
            )
            assets.append(asset)
        
        if not assets:
            raise ValueError("No assets found in Excel file")
        
        portfolio = Portfolio(assets=assets, cash_available=cash_available)
        return portfolio
    
    def _adjust_column_widths(self, sheet) -> None:
        """Adjust column widths based on content.
        
        Args:
            sheet: openpyxl worksheet
        """
        for column_cells in sheet.columns:
            max_length = 0
            # Get column letter from first non-merged cell
            column_letter = None
            for cell in column_cells:
                # Skip merged cells
                if hasattr(cell, 'column_letter'):
                    if column_letter is None:
                        column_letter = cell.column_letter
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = max(adjusted_width, 12)
    
    def write_result(
        self, result: RebalancingResult, filepath: str, sheet_name: str = "Rebalancing"
    ) -> None:
        """Write rebalancing results to Excel file.
        
        Args:
            result: RebalancingResult to write
            filepath: Path to output Excel file
            sheet_name: Name of sheet to create
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)
        red_font = Font(color="9C0006")  # Red for negative cash flows
        
        # Title
        sheet["A1"] = "Portfolio Rebalancing Results"
        sheet["A1"].font = title_font
        sheet.merge_cells("A1:H1")
        
        # Summary section
        row = 3
        sheet[f"A{row}"] = "Summary"
        sheet[f"A{row}"].font = section_font
        
        row += 1
        cash_available = result.metadata.get("cash_available", 0.0)
        if cash_available > 0:
            sheet[f"A{row}"] = "Cash Available to Invest:"
            sheet[f"B{row}"] = cash_available
            sheet[f"B{row}"].number_format = '€#,##0.00'
            row += 1
        
        sheet[f"A{row}"] = "Total Portfolio Value Before:"
        sheet[f"B{row}"] = result.total_value_before
        sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Total Portfolio Value After:"
        sheet[f"B{row}"] = result.total_value_after
        sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Cash Flow Required:"
        # Cash flow: negative = need to add money (red), positive = money freed (black)
        if result.cash_flow < 0:
            sheet[f"B{row}"] = abs(result.cash_flow)  # Store as positive
            sheet[f"B{row}"].number_format = '-€#,##0.00'  # Display with minus
            sheet[f"B{row}"].font = red_font
        else:
            sheet[f"B{row}"] = result.cash_flow
            sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Total Cash In (from sales):"
        sheet[f"B{row}"] = result.total_cash_in
        sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Total Cash Out (for purchases):"
        # total_cash_out is stored as positive, show as negative in red
        if result.total_cash_out > 0:
            sheet[f"B{row}"] = result.total_cash_out  # Store as positive
            sheet[f"B{row}"].number_format = '-€#,##0.00'  # Display with minus
            sheet[f"B{row}"].font = red_font
        else:
            sheet[f"B{row}"] = 0.0
            sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Total Tax Paid:"
        # Tax is stored as positive, show as negative in red
        if result.total_tax_paid > 0:
            sheet[f"B{row}"] = result.total_tax_paid  # Store as positive
            sheet[f"B{row}"].number_format = '-€#,##0.00'  # Display with minus
            sheet[f"B{row}"].font = red_font
        else:
            sheet[f"B{row}"] = 0.0
            sheet[f"B{row}"].number_format = '€#,##0.00'
        
        # Commission summary section
        row += 1
        sheet[f"A{row}"] = "Total Commission on Purchases:"
        # Commission is a cost, show as negative in red
        if result.total_commission_buy > 0:
            sheet[f"B{row}"] = result.total_commission_buy  # Store as positive
            sheet[f"B{row}"].number_format = '-€#,##0.00'  # Display with minus
            sheet[f"B{row}"].font = red_font
        else:
            sheet[f"B{row}"] = 0.0
            sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Total Commission on Sales:"
        # Commission is a cost, show as negative in red
        if result.total_commission_sell > 0:
            sheet[f"B{row}"] = result.total_commission_sell  # Store as positive
            sheet[f"B{row}"].number_format = '-€#,##0.00'  # Display with minus
            sheet[f"B{row}"].font = red_font
        else:
            sheet[f"B{row}"] = 0.0
            sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Total Commissions:"
        # Total commission is a cost, show as negative in red
        if result.total_commission > 0:
            sheet[f"B{row}"] = result.total_commission  # Store as positive
            sheet[f"B{row}"].number_format = '-€#,##0.00'  # Display with minus
            sheet[f"B{row}"].font = red_font
        else:
            sheet[f"B{row}"] = 0.0
            sheet[f"B{row}"].number_format = '€#,##0.00'
        
        row += 1
        sheet[f"A{row}"] = "Max Deviation from Target:"
        sheet[f"B{row}"] = result.max_deviation
        sheet[f"B{row}"].number_format = '0.00%'
        
        row += 1
        sheet[f"A{row}"] = "Number of Operations:"
        sheet[f"B{row}"] = result.num_operations
        
        # Operations section - simplified (removed Current Qty, New Qty, Target Weight)
        row += 3
        sheet[f"A{row}"] = "Operations Needed"
        sheet[f"A{row}"].font = section_font
        
        row += 1
        headers = [
            "Symbol",
            "Action",
            "Quantity Change",
            "Value Change (€)",
            "Current Weight",
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Asset rows
        for asset in result.assets:
            row += 1
            action = "BUY" if asset.delta_quantity > 0 else ("SELL" if asset.delta_quantity < 0 else "HOLD")
            
            sheet[f"A{row}"] = asset.symbol
            sheet[f"B{row}"] = action
            sheet[f"C{row}"] = asset.delta_quantity
            sheet[f"C{row}"].number_format = '0.00'
            
            # Value Change logic (CORRECTED):
            # delta_value > 0 = portfolio value increases (BUY) → positive value
            # delta_value < 0 = portfolio value decreases (SELL) → negative value, display as positive with minus sign
            if asset.delta_value < 0:  # SELL: portfolio value decreases (negative)
                sheet[f"D{row}"] = abs(asset.delta_value)  # Store absolute value
                sheet[f"D{row}"].number_format = '-€#,##0.00'  # Display with minus
                sheet[f"D{row}"].font = red_font
            else:  # BUY or HOLD: portfolio value increases or stays same (positive/zero)
                sheet[f"D{row}"] = asset.delta_value  # Store as-is
                sheet[f"D{row}"].number_format = '€#,##0.00'
            
            sheet[f"E{row}"] = asset.current_weight
            sheet[f"E{row}"].number_format = '0.00%'
        
        # Final Portfolio section
        row += 3
        sheet[f"A{row}"] = "Final Portfolio Weights"
        sheet[f"A{row}"].font = section_font
        
        row += 1
        final_headers = [
            "Symbol",
            "Final Quantity",
            "Price",
            "Final Value (€)",
            "Final Weight",
            "Target Weight",
            "Deviation",
        ]
        
        for col_idx, header in enumerate(final_headers, start=1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Final portfolio rows
        for asset in result.assets:
            row += 1
            final_qty = asset.quantity + asset.delta_quantity
            final_value = final_qty * asset.price
            final_weight = final_value / result.total_value_after if result.total_value_after > 0 else 0
            deviation = final_weight - asset.target_weight
            
            sheet[f"A{row}"] = asset.symbol
            sheet[f"B{row}"] = final_qty
            sheet[f"B{row}"].number_format = '0.00'
            sheet[f"C{row}"] = asset.price
            sheet[f"C{row}"].number_format = '€#,##0.00'
            sheet[f"D{row}"] = final_value
            sheet[f"D{row}"].number_format = '€#,##0.00'
            sheet[f"E{row}"] = final_weight
            sheet[f"E{row}"].number_format = '0.00%'
            sheet[f"F{row}"] = asset.target_weight
            sheet[f"F{row}"].number_format = '0.00%'
            sheet[f"G{row}"] = deviation
            sheet[f"G{row}"].number_format = '0.00%'
            
            # Deviation color coding (no bold, just colors):
            # Green: |deviation| < 0.25%
            # Yellow: 0.25% <= |deviation| < 1%
            # Red: |deviation| >= 1%
            abs_deviation = abs(deviation)
            if abs_deviation < 0.0025:  # < 0.25%
                sheet[f"G{row}"].font = Font(color="006100")  # Green (no bold)
            elif abs_deviation < 0.01:  # 0.25% to 1%
                sheet[f"G{row}"].font = Font(color="9C6500")  # Yellow (no bold)
            else:  # >= 1%
                sheet[f"G{row}"].font = Font(color="9C0006")  # Red (no bold)
        
        # Total row
        row += 1
        sheet[f"A{row}"] = "TOTAL"
        sheet[f"A{row}"].font = Font(bold=True)
        sheet[f"D{row}"] = result.total_value_after
        sheet[f"D{row}"].number_format = '€#,##0.00'
        sheet[f"D{row}"].font = Font(bold=True)
        
        # Sum of final weights (should be 100%)
        total_weight = sum(
            (asset.quantity + asset.delta_quantity) * asset.price / result.total_value_after
            for asset in result.assets
        ) if result.total_value_after > 0 else 0
        sheet[f"E{row}"] = total_weight
        sheet[f"E{row}"].number_format = '0.00%'
        sheet[f"E{row}"].font = Font(bold=True)
        
        # Adjust column widths
        self._adjust_column_widths(sheet)
        
        # Save workbook
        try:
            workbook.save(filepath)
        except Exception as e:
            raise IOError(f"Failed to save Excel file: {e}")
    
    def write_portfolio(
        self, portfolio: Portfolio, filepath: str, sheet_name: str = "Portfolio"
    ) -> None:
        """Write portfolio to Excel file (template format).
        
        Args:
            portfolio: Portfolio to write
            filepath: Path to output Excel file
            sheet_name: Name of sheet to create
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        
        # Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        label_font = Font(bold=True, size=11)
        
        # Cash Available section
        sheet["A1"] = "Cash Available to Invest:"
        sheet["A1"].font = label_font
        sheet["B1"] = portfolio.cash_available
        sheet["B1"].number_format = '€#,##0.00'
        
        # Empty row for separation
        # Row 2 is empty
        
        # Headers in row 3 - now with all commission columns
        headers = [
            "Symbol",
            "Quantity",
            "Price",
            "Avg Cost",
            "Tax Rate",
            "Target Weight",
            "Comm Buy Fixed",
            "Comm Buy %",
            "Comm Buy Min",
            "Comm Buy Max",
            "Comm Sell Fixed",
            "Comm Sell %",
            "Comm Sell Min",
            "Comm Sell Max",
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Asset rows starting from row 4
        for row_idx, asset in enumerate(portfolio.assets, start=4):
            sheet[f"A{row_idx}"] = asset.symbol
            sheet[f"B{row_idx}"] = asset.quantity
            sheet[f"B{row_idx}"].number_format = '0.00'
            sheet[f"C{row_idx}"] = asset.price
            sheet[f"C{row_idx}"].number_format = '€#,##0.00'
            sheet[f"D{row_idx}"] = asset.avg_cost
            sheet[f"D{row_idx}"].number_format = '€#,##0.00'
            sheet[f"E{row_idx}"] = asset.tax_rate
            sheet[f"E{row_idx}"].number_format = '0.00%'
            sheet[f"F{row_idx}"] = asset.target_weight
            sheet[f"F{row_idx}"].number_format = '0.00%'
            
            # Commission fields
            sheet[f"G{row_idx}"] = asset.commission_buy_fixed
            sheet[f"G{row_idx}"].number_format = '€#,##0.00'
            sheet[f"H{row_idx}"] = asset.commission_buy_percent
            sheet[f"H{row_idx}"].number_format = '0.000%'
            sheet[f"I{row_idx}"] = asset.commission_buy_min
            sheet[f"I{row_idx}"].number_format = '€#,##0.00'
            sheet[f"J{row_idx}"] = asset.commission_buy_max
            sheet[f"J{row_idx}"].number_format = '€#,##0.00'
            
            sheet[f"K{row_idx}"] = asset.commission_sell_fixed
            sheet[f"K{row_idx}"].number_format = '€#,##0.00'
            sheet[f"L{row_idx}"] = asset.commission_sell_percent
            sheet[f"L{row_idx}"].number_format = '0.000%'
            sheet[f"M{row_idx}"] = asset.commission_sell_min
            sheet[f"M{row_idx}"].number_format = '€#,##0.00'
            sheet[f"N{row_idx}"] = asset.commission_sell_max
            sheet[f"N{row_idx}"].number_format = '€#,##0.00'
        
        # Adjust column widths
        self._adjust_column_widths(sheet)
        
        # Save workbook
        try:
            workbook.save(filepath)
        except Exception as e:
            raise IOError(f"Failed to save Excel file: {e}")
